"""
Vista de Sincronización: exporta los datos de un mes específico a un archivo
.json para transferirlo (USB, correo) a la otra PC. La importación/fusión
de esos archivos se implementa en un paso posterior.
"""

import customtkinter as ctk
from tkinter import messagebox, filedialog
import json
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from gui import estilos
from database.consultas import (
    listar_meses_disponibles,
    exportar_datos_mes,
    procesar_importacion,
    registrar_log_sincronizacion,
    obtener_estudiante_por_id,
    insertar_atencion_con_id,
    listar_usuarios_por_ids,
)

from config import obtener_origen_pc

MESES_ES = {
    "01": "Enero", "02": "Febrero", "03": "Marzo", "04": "Abril",
    "05": "Mayo", "06": "Junio", "07": "Julio", "08": "Agosto",
    "09": "Septiembre", "10": "Octubre", "11": "Noviembre", "12": "Diciembre",
}

def _nombre_mes(mes_iso):
    anio, mes = mes_iso.split("-")
    return f"{MESES_ES.get(mes, mes)} {anio}"

class VistaSincronizacion(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master, fg_color="transparent")
        self.grid_columnconfigure(0, weight=1)
        self._crear_widgets()

    def _crear_widgets(self):
        titulo = ctk.CTkLabel(self, text="Sincronización", font=estilos.fuente_titulo())
        titulo.grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 5))

        subtitulo = ctk.CTkLabel(
            self, text=f"Esta PC está identificada como: {obtener_origen_pc()}",
            text_color=estilos.COLOR_TEXTO_GRIS
        )
        subtitulo.grid(row=1, column=0, columnspan=2, sticky="w", pady=(0, 20))

        # Dos columnas de igual ancho, para que las tarjetas queden lado a lado
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)

        # --- Tarjeta de exportación (columna izquierda) ---
        tarjeta = ctk.CTkFrame(self)
        tarjeta.grid(row=2, column=0, sticky="new", padx=(0, 10))
        tarjeta.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            tarjeta, text="Exportar datos", font=estilos.fuente_seccion_pequena(), anchor="w"
        ).grid(row=0, column=0, padx=15, pady=(15, 5), sticky="w")

        ctk.CTkLabel(
            tarjeta,
            text="Genera un archivo con los estudiantes y atenciones del mes\n"
                "elegido, para transferirlo a la otra PC (USB o correo).",
            text_color=estilos.COLOR_TEXTO_GRIS, anchor="w", justify="left"
        ).grid(row=1, column=0, padx=15, pady=(0, 15), sticky="w")

        ctk.CTkLabel(tarjeta, text="Mes a exportar", anchor="w").grid(
            row=2, column=0, padx=15, sticky="w"
        )

        meses = listar_meses_disponibles()
        etiquetas_meses = [_nombre_mes(m) for m in meses] or ["Sin datos disponibles"]
        self._mapa_mes = dict(zip(etiquetas_meses, meses)) if meses else {}

        self.combo_mes = ctk.CTkComboBox(
            tarjeta, values=etiquetas_meses, state="readonly" if meses else "disabled"
        )
        self.combo_mes.set(etiquetas_meses[0])
        self.combo_mes.grid(row=3, column=0, padx=15, pady=(0, 15), sticky="ew")

        self.btn_exportar = ctk.CTkButton(
            tarjeta, text="Exportar mes seleccionado (.json)", command=self._exportar,
            state="normal" if meses else "disabled"
        )
        self.btn_exportar.grid(row=4, column=0, padx=15, pady=(0, 10), sticky="ew")

        self.btn_exportar_excel = ctk.CTkButton(
            tarjeta, text="Exportar a Excel (.xlsx)", command=self._exportar_excel,
            fg_color=estilos.COLOR_AMARILLO, hover_color=estilos.COLOR_AMARILLO_HOVER,
            state="normal" if meses else "disabled"
        )
        self.btn_exportar_excel.grid(row=5, column=0, padx=15, pady=(0, 20), sticky="ew")

        # --- Tarjeta de importación (columna derecha) ---
        tarjeta_importar = ctk.CTkFrame(self)
        tarjeta_importar.grid(row=2, column=1, sticky="new", padx=(10, 0))
        tarjeta_importar.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            tarjeta_importar, text="Importar datos", font=estilos.fuente_seccion_pequena(), anchor="w"
        ).grid(row=0, column=0, padx=15, pady=(15, 5), sticky="w")

        ctk.CTkLabel(
            tarjeta_importar,
            text="Selecciona el archivo .json exportado desde la otra\n"
                "PC para fusionarlo con los datos de esta computadora.",
            text_color=estilos.COLOR_TEXTO_GRIS, anchor="w", justify="left"
        ).grid(row=1, column=0, padx=15, pady=(0, 15), sticky="w")

        ctk.CTkButton(
            tarjeta_importar, text="Seleccionar archivo e importar", command=self._importar
        ).grid(row=2, column=0, padx=15, pady=(0, 20), sticky="ew")

    def _exportar(self):
        etiqueta_mes = self.combo_mes.get()
        mes_iso = self._mapa_mes.get(etiqueta_mes)
        if not mes_iso:
            messagebox.showwarning("Atención", "No hay un mes válido seleccionado.")
            return

        estudiantes, atenciones = exportar_datos_mes(mes_iso)

        if not atenciones:
            messagebox.showinfo("Sin datos", f"No hay atenciones registradas en {etiqueta_mes}.")
            return

        paquete = {
            "origen_pc": obtener_origen_pc(),
            "mes": mes_iso,
            "fecha_exportacion": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "estudiantes": estudiantes,
            "atenciones": atenciones,
        }

        nombre_sugerido = f"enfermeria_{obtener_origen_pc()}_{mes_iso}.json"
        ruta_destino = filedialog.asksaveasfilename(
            title="Guardar exportación",
            initialfile=nombre_sugerido,
            defaultextension=".json",
            filetypes=[("Archivo JSON", "*.json")]
        )
        if not ruta_destino:
            return  # el usuario canceló el diálogo

        try:
            with open(ruta_destino, "w", encoding="utf-8") as archivo:
                json.dump(paquete, archivo, ensure_ascii=False, indent=2)
        except Exception as error:
            messagebox.showerror("Error al exportar", f"No se pudo guardar el archivo:\n{error}")
            return

        messagebox.showinfo(
            "Exportación exitosa",
            f"Se exportaron {len(atenciones)} atención(es) de {len(estudiantes)} estudiante(s)\n"
            f"correspondientes a {etiqueta_mes}.\n\nArchivo guardado en:\n{ruta_destino}"
        )
        
    def _exportar_excel(self):
        etiqueta_mes = self.combo_mes.get()
        mes_iso = self._mapa_mes.get(etiqueta_mes)
        if not mes_iso:
            messagebox.showwarning("Atención", "No hay un mes válido seleccionado.")
            return

        estudiantes, atenciones = exportar_datos_mes(mes_iso)
        if not atenciones:
            messagebox.showinfo("Sin datos", f"No hay atenciones registradas en {etiqueta_mes}.")
            return

        mapa_estudiantes = {e["id"]: e for e in estudiantes}
        ids_enfermeras = {a.get("enfermera_responsable") for a in atenciones}
        mapa_enfermeras = listar_usuarios_por_ids(ids_enfermeras)

        atenciones_ordenadas = sorted(atenciones, key=lambda a: (a["fecha"], a["hora_llegada"]))

        wb = Workbook()
        ws = wb.active
        ws.title = "Atenciones"

        encabezados = [
            "Fecha", "Hora Llegada", "Hora Salida", "Nombre", "Curso", "Paralelo", "Sexo",
            "Saturación (%)", "Temperatura (°C)", "Frecuencia Cardíaca (lpm)",
            "Diagnóstico", "Recomendación", "Enfermera Responsable", "Origen PC",
        ]
        ws.append(encabezados)

        fill_encabezado = PatternFill(start_color="1C1C1C", end_color="1C1C1C", fill_type="solid")
        font_encabezado = Font(color="FFC72C", bold=True)
        for celda in ws[1]:
            celda.fill = fill_encabezado
            celda.font = font_encabezado
            celda.alignment = Alignment(horizontal="center", vertical="center")

        for at in atenciones_ordenadas:
            est = mapa_estudiantes.get(at["estudiante_id"], {})
            nombre_enfermera = mapa_enfermeras.get(at.get("enfermera_responsable"), "—")
            ws.append([
                at["fecha"], at["hora_llegada"], at.get("hora_salida") or "—",
                est.get("nombre", "—"), est.get("curso", "—"), est.get("paralelo") or "—",
                est.get("sexo", "—"),
                at.get("saturacion"), at.get("temperatura"), at.get("frecuencia_cardiaca"),
                at.get("diagnostico") or "—", at.get("recomendacion") or "—",
                nombre_enfermera, at.get("origen_pc") or "—",
            ])

        anchos = [12, 12, 12, 22, 16, 12, 12, 14, 16, 18, 24, 30, 20, 10]
        for i, ancho in enumerate(anchos, start=1):
            ws.column_dimensions[get_column_letter(i)].width = ancho

        ws.freeze_panes = "A2"  # deja fija la fila de encabezados al hacer scroll

        nombre_sugerido = f"enfermeria_{obtener_origen_pc()}_{mes_iso}.xlsx"
        ruta_destino = filedialog.asksaveasfilename(
            title="Guardar Excel",
            initialfile=nombre_sugerido,
            defaultextension=".xlsx",
            filetypes=[("Libro de Excel", "*.xlsx")]
        )
        if not ruta_destino:
            return

        try:
            wb.save(ruta_destino)
        except Exception as error:
            messagebox.showerror("Error al exportar", f"No se pudo guardar el archivo:\n{error}")
            return

        messagebox.showinfo(
            "Exportación exitosa",
            f"Se exportaron {len(atenciones)} atención(es) a Excel.\n\nArchivo guardado en:\n{ruta_destino}"
        )
    # ------------------------------------------------------------
    # Importación
    # ------------------------------------------------------------
    def _importar(self):
        ruta = filedialog.askopenfilename(
            title="Selecciona el archivo exportado",
            filetypes=[("Archivo JSON", "*.json")]
        )
        if not ruta:
            return

        try:
            with open(ruta, "r", encoding="utf-8") as archivo:
                paquete = json.load(archivo)
        except Exception as error:
            messagebox.showerror("Error al leer archivo", f"No se pudo leer el archivo:\n{error}")
            return

        if "estudiantes" not in paquete or "atenciones" not in paquete:
            messagebox.showerror("Archivo inválido", "Este archivo no tiene el formato esperado de exportación.")
            return

        if paquete.get("origen_pc") == obtener_origen_pc():
            confirmar = messagebox.askyesno(
                "Mismo origen",
                f"Este archivo fue exportado desde ESTA MISMA PC ({obtener_origen_pc()}).\n"
                "Normalmente solo deberías importar archivos generados por la OTRA PC.\n\n"
                "¿Deseas continuar de todas formas?"
            )
            if not confirmar:
                return

        resumen, para_revisar = procesar_importacion(paquete)

        registrar_log_sincronizacion(
            registros_nuevos=resumen["atenciones_nuevas"],
            duplicados_detectados=resumen["atenciones_para_revisar"],
            notas=f"Importado desde {paquete.get('origen_pc', '?')} — mes {paquete.get('mes', '?')}"
        )

        mensaje = (
            f"Estudiantes nuevos creados: {resumen['estudiantes_nuevos']}\n"
            f"Estudiantes ya existentes (reutilizados): {resumen['estudiantes_reutilizados']}\n"
            f"Atenciones nuevas importadas: {resumen['atenciones_nuevas']}\n"
            f"Atenciones ya existentes (omitidas): {resumen['atenciones_ya_existentes']}\n"
            f"Posibles duplicados detectados: {resumen['atenciones_para_revisar']}"
        )
        messagebox.showinfo("Importación completada", mensaje)

        if para_revisar:
            self._mostrar_revision_duplicados(para_revisar)

    def _mostrar_revision_duplicados(self, para_revisar):
        ventana = ctk.CTkToplevel(self)
        ventana.title("Revisar posibles duplicados")
        ventana.geometry("620x520")
        ventana.grab_set()

        ctk.CTkLabel(
            ventana,
            text=f"Se encontraron {len(para_revisar)} posible(s) duplicado(s).\n"
                "Revisa cada caso: si es la misma visita registrada por error en ambas PCs,\n"
                "descártala. Si son visitas distintas que coincidieron en horario, impórtala.",
            justify="left"
        ).pack(padx=15, pady=15, anchor="w")

        frame_lista = ctk.CTkScrollableFrame(ventana)
        frame_lista.pack(fill="both", expand=True, padx=15, pady=(0, 10))

        self._decisiones_duplicados = []

        for item in para_revisar:
            at = item["atencion_importada"]
            estudiante = obtener_estudiante_por_id(item["estudiante_id_local"])
            nombre_estudiante = estudiante[1] if estudiante else "Desconocido"

            tarjeta = ctk.CTkFrame(frame_lista)
            tarjeta.pack(fill="x", pady=6, padx=4)

            ctk.CTkLabel(
                tarjeta, text=f"{nombre_estudiante} — {at['fecha']} {at['hora_llegada']}",
                font=estilos.fuente_etiqueta(), anchor="w"
            ).pack(fill="x", padx=10, pady=(10, 0))

            ctk.CTkLabel(
                tarjeta,
                text=f"Diagnóstico importado: {at.get('diagnostico') or '—'}  (origen: {at.get('origen_pc')})",
                text_color=estilos.COLOR_TEXTO_GRIS, anchor="w"
            ).pack(fill="x", padx=10)

            for similar in item["similares_existentes"]:
                ctk.CTkLabel(
                    tarjeta,
                    text=f"⚠ Ya existe: {similar['hora_llegada']} — {similar['diagnostico'] or '—'} "
                        f"(origen: {similar['origen_pc']})",
                    text_color=estilos.COLOR_AMARILLO, anchor="w"
                ).pack(fill="x", padx=10)

            var_decision = ctk.StringVar(value="descartar")
            frame_opciones = ctk.CTkFrame(tarjeta, fg_color="transparent")
            frame_opciones.pack(fill="x", padx=10, pady=(5, 10))
            ctk.CTkRadioButton(
                frame_opciones, text="Descartar (es el mismo registro)",
                variable=var_decision, value="descartar"
            ).pack(side="left", padx=(0, 15))
            ctk.CTkRadioButton(
                frame_opciones, text="Importar de todas formas (es distinto)",
                variable=var_decision, value="importar"
            ).pack(side="left")

            self._decisiones_duplicados.append({"item": item, "var": var_decision})

        ctk.CTkButton(
            ventana, text="Aplicar decisiones", command=lambda: self._aplicar_decisiones(ventana)
        ).pack(pady=(0, 15))

    def _aplicar_decisiones(self, ventana):
        importados = 0
        for decision in self._decisiones_duplicados:
            if decision["var"].get() == "importar":
                item = decision["item"]
                insertar_atencion_con_id(item["atencion_importada"], item["estudiante_id_local"])
                importados += 1

        messagebox.showinfo("Listo", f"Se importaron {importados} registro(s) adicionales que marcaste como distintos.")
        ventana.destroy()